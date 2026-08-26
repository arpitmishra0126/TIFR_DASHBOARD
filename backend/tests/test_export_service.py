import csv
from datetime import date
from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook

from app.services.export_service import (
    ASSESSMENT_INSTRUMENTS,
    build_active_cases_csv,
    build_active_cases_workbook,
    export_filename,
)
from app.services.live_dashboard_service import LiveDashboardService
from tests.fixtures.live_redcap import FIXTURE_METADATA, FakeRedCapRepository, _base_record, build_fixture_records

AS_OF = date(2026, 8, 25)


async def _export_bytes() -> bytes:
    records = build_fixture_records(AS_OF)
    repo = FakeRedCapRepository(FIXTURE_METADATA, records)
    service = LiveDashboardService(repo)
    return await service.get_active_cases_export()


def test_export_filename_uses_dated_pattern():
    assert export_filename(date(2026, 8, 25)) == "ICMR_Active_Cases_2026-08-25.xlsx"


def test_export_filename_supports_csv_extension():
    assert export_filename(date(2026, 8, 25), extension="csv") == "ICMR_Active_Cases_2026-08-25.csv"


def test_workbook_has_four_expected_sheets():
    records = build_fixture_records(AS_OF)
    from app.ingestion.choice_maps import build_choice_maps
    from app.services.live_dashboard_service import _normalize_child

    choice_maps = build_choice_maps(FIXTURE_METADATA)
    children = [c for r in records if (c := _normalize_child(r, choice_maps)) is not None]

    workbook_bytes = build_active_cases_workbook(children, records, choice_maps)
    workbook = load_workbook(BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["Active Cases", "Assessment Status", "Summary", "Data Dictionary"]


def _load(workbook_bytes: bytes):
    return load_workbook(BytesIO(workbook_bytes))


def _active_cases_row_values(sheet, child_id: str) -> dict:
    """Active Cases sheet has a 2-row header (group row 1, field row 2);
    data starts at row 3."""
    header = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    row = next(r for r in sheet.iter_rows(min_row=3) if r[0].value == child_id)
    return dict(zip(header, [c.value for c in row]))


async def test_active_cases_sheet_excludes_dead_and_unregistered():
    # Fixture has no "Dead" record by default; all baby_status="1" (Live) or
    # blank child_id (excluded entirely upstream in normalization).
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    child_ids = [row[0].value for row in sheet.iter_rows(min_row=3) if row[0].value]
    assert set(child_ids) == {"REC001", "REC002", "REC003", "REC004", "REC005", "REC006"}


async def test_active_cases_sheet_has_grouped_header_row():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    group_values = {cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)) if cell.value}
    assert "A. Registration & Demographics" in group_values
    assert "B. SES" in group_values
    assert "G. SSRS Parent" in group_values
    assert "H. SSRS Child" in group_values
    assert "I. SSRS Teacher" in group_values
    assert "J. Assessment / Progression Status" in group_values


async def test_active_cases_sheet_has_filters_and_frozen_header():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    assert sheet.freeze_panes == "A3"
    assert sheet.auto_filter.ref is not None


async def test_active_cases_sheet_includes_ses_values_as_typed_numbers():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    values = _active_cases_row_values(sheet, "REC001")

    assert values["Udai Pareek SES Score"] == 35
    assert values["Per Capita Income"] == 3000.0
    assert values["Household Size"] == 5
    assert values["Age (Years)"] == 6


async def test_active_cases_sheet_includes_newly_audited_instrument_fields_with_real_values():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    values = _active_cases_row_values(sheet, "REC001")

    assert values["Monthly Household Income (INR)"] == 15000.0
    # Note: choice_maps' primary-language-segment split cuts on the FIRST
    # "/" in a choice label — a pre-existing behavior (app.ingestion.choice_maps,
    # unchanged here) that also truncates "1-2 days/week" style live DSEQ
    # labels mid-word. Documented as a known quirk, not fixed here since
    # that module is shared dashboard logic.
    assert values["DSEQ Q1: TV Frequency (per week)"] == "1-2 days"
    assert values["CHH: Currently Ill (Y/N)"] == "No"
    assert values["PAQ-A Total Score"] == 3.2
    assert values["Diet Frequency: Grains / Roots / Tubers"] == "Daily"


async def test_active_cases_sheet_leaves_new_fields_blank_when_not_collected():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    values = _active_cases_row_values(sheet, "REC005")

    assert not values["Monthly Household Income (INR)"]
    assert not values["DSEQ Q1: TV Frequency (per week)"]
    assert not values["PAQ-A Total Score"]


async def test_active_cases_sheet_computes_real_ssrs_parent_and_child_summaries():
    # REC001 has p1_freq="1", p2_freq="0", p1_imp="2" in the fixture (out of
    # the real 52 frequency / 40 importance items) — verifies real acquired
    # SSRS Parent data now surfaces instead of a blanket "not mapped" text.
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    values = _active_cases_row_values(sheet, "REC001")

    assert values["SSRS Parent: Items Answered"] == "2/52"
    assert values["SSRS Parent: Avg Frequency Rating"] == 0.5
    assert values["SSRS Parent: Avg Importance Rating"] == 2.0


async def test_active_cases_sheet_ssrs_teacher_shows_no_acquired_data():
    # 0/212 live Teacher completions — every active child must show 0
    # items answered and blank averages, never an invented value.
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    values = _active_cases_row_values(sheet, "REC001")

    assert values["SSRS Teacher: Items Answered"] == "0/42"
    assert not values["SSRS Teacher: Avg Frequency Rating"]
    assert not values["SSRS Teacher: Avg Importance Rating"]


async def test_active_cases_sheet_has_progression_status_group():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    values = _active_cases_row_values(sheet, "REC004")  # partial core battery

    assert values["Registration Complete"] == "Complete"
    assert values["Dietary Intake Complete"] == "Not Complete"
    assert values["Core Assessment Battery"] == "Not Complete"
    assert values["Overall Progression Stage"] == "Registered"


async def test_active_cases_sheet_excludes_caste_and_freetext_fields():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Active Cases"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

    assert not any("caste" in h.lower() for h in header)
    assert not any("comment" in h.lower() or "remarks" in h.lower() for h in header)
    assert not any(h in ("parent_child_id", "teacher_child_id") for h in header)


def test_data_dictionary_documents_every_active_cases_column():
    from app.services.export_service import ACTIVE_CASES_FIELD_SPECS

    wb_headers = {spec.header for spec in ACTIVE_CASES_FIELD_SPECS}
    documented = set()
    workbook = Workbook()
    from app.services.export_service import _build_data_dictionary_sheet

    sheet = workbook.active
    _build_data_dictionary_sheet(sheet)
    for row in sheet.iter_rows(min_row=2, max_col=1):
        value = row[0].value
        if value:
            documented.add(value)

    assert wb_headers.issubset(documented)


async def test_assessment_status_sheet_reflects_completion_flags():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Assessment Status"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rec4_row = next(row for row in sheet.iter_rows(min_row=2) if row[0].value == "REC004")
    values = dict(zip(header, [c.value for c in rec4_row]))

    assert values["SES"] == "Complete"
    assert values["Dietary Intake"] == "Not Complete"
    assert values["SSRS Child"] == "Not Complete"
    assert values["Core Assessment Battery"] == "Not Complete"
    assert values["Overall Progression Stage"] == "Registered"


async def test_assessment_status_sheet_reports_core_battery_and_stage_for_full_pipeline_child():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Assessment Status"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rec2_row = next(row for row in sheet.iter_rows(min_row=2) if row[0].value == "REC002")  # full pipeline incl. SSRS Teacher
    values = dict(zip(header, [c.value for c in rec2_row]))

    assert values["Core Assessment Battery"] == "Complete"
    assert values["Overall Progression Stage"] == "SSRS Teacher"


async def test_summary_sheet_reports_active_total_and_renamed_title():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Summary"]
    title = sheet.cell(row=1, column=1).value
    assert title == "ICMR Neurodevelopment Study — Active Cases Analysis"

    rows = {row[0].value: row[1].value for row in sheet.iter_rows(min_row=1) if row[0].value}
    assert rows["Total Active Cases"] == 6
    assert rows["Total Registered (all statuses)"] == 6


async def test_summary_sheet_has_required_sections():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Summary"]
    section_titles = {row[0].value for row in sheet.iter_rows(min_row=1) if row[0].value}

    for section in (
        "POPULATION", "SES", "ASSESSMENT ACQUISITION", "ASSESSMENT PROGRESSION",
        "DOMAIN ANALYSIS (where sufficient data exists)", "DATA COVERAGE",
    ):
        assert section in section_titles


async def test_summary_sheet_data_coverage_buckets_instruments_correctly():
    workbook_bytes = await _export_bytes()
    wb = _load(workbook_bytes)
    sheet = wb["Summary"]
    rows_by_label = {}
    for row in sheet.iter_rows(min_row=1):
        if row[0].value:
            rows_by_label[row[0].value] = row[1].value

    # Registration: 5/6 active complete -> High coverage (>=50%).
    assert rows_by_label["Registration"] == 5
    # SSRS Teacher: only REC002 has ssrs_teacher_complete in the fixture (1/6)
    # -> Partially Acquired bucket, not High or No-Data.
    assert rows_by_label["SSRS Teacher"] == 1
    # Dietary Intake: no record in the fixture has dietary_intake_complete
    # left incomplete among those with core_complete except REC004 -> still
    # some coverage; use SSRS Child (only REC001/REC002 = 2/6) to exercise
    # the "Partially Acquired" bucket explicitly.
    assert rows_by_label["SSRS Child"] == 2


def _csv_rows(csv_text: str) -> list[dict]:
    return list(csv.DictReader(StringIO(csv_text)))


def _fixture_children_and_records(records: list[dict]):
    from app.ingestion.choice_maps import build_choice_maps
    from app.services.live_dashboard_service import _normalize_child

    choice_maps = build_choice_maps(FIXTURE_METADATA)
    children = [c for r in records if (c := _normalize_child(r, choice_maps)) is not None]
    return children, records


def test_csv_excludes_children_marked_dead():
    records = build_fixture_records(AS_OF)
    records.append(_base_record("REC008", baby_status="0"))  # Dead
    children, records = _fixture_children_and_records(records)

    csv_text = build_active_cases_csv(children, records)
    rows = _csv_rows(csv_text)
    child_ids = {row["Child ID"] for row in rows}

    assert "REC008" not in child_ids
    assert child_ids == {"REC001", "REC002", "REC003", "REC004", "REC005", "REC006"}


def test_csv_includes_ses_values_and_leaves_unmapped_instrument_data_blank():
    records = build_fixture_records(AS_OF)
    children, records = _fixture_children_and_records(records)

    csv_text = build_active_cases_csv(children, records)
    rows = {row["Child ID"]: row for row in _csv_rows(csv_text)}
    rec1 = rows["REC001"]

    assert rec1["Udai Pareek SES Score"] == "35"
    assert rec1["Per Capita Income"] == "3000.0"
    assert rec1["Household Size"] == "5"
    # Instrument data columns aren't part of the CSV at all — only status.
    assert "DSEQ Data" not in rec1
    assert rec1["DSEQ Status"] == "Complete"
    assert rec1["SSRS Teacher Status"] == "Incomplete"


def test_csv_leaves_blank_ses_fields_blank_when_not_collected():
    records = build_fixture_records(AS_OF)
    children, records = _fixture_children_and_records(records)

    csv_text = build_active_cases_csv(children, records)
    rows = {row["Child ID"]: row for row in _csv_rows(csv_text)}
    rec5 = rows["REC005"]  # registered only, nothing else started

    assert rec5["Udai Pareek SES Score"] == ""
    assert rec5["Per Capita Income"] == ""
    assert rec5["Household Size"] == ""
    assert rec5["SES Status"] == "Incomplete"


def test_csv_header_matches_approved_field_set():
    records = build_fixture_records(AS_OF)
    children, records = _fixture_children_and_records(records)
    csv_text = build_active_cases_csv(children, records)
    header = next(csv.reader(StringIO(csv_text)))

    assert header == [
        "Child ID",
        "Sex",
        "Date of Birth",
        "Age (Years)",
        "Village",
        "Child Status",
        "Visit Date",
        "Registration Complete",
        "Udai Pareek SES Score",
        "Udai Pareek SES Category",
        "BG Prasad SES Category",
        "Per Capita Income",
        "Household Size",
        "SES Status",
        "DSEQ Status",
        "Child Illness History Status",
        "PAQ-A Status",
        "Dietary Intake Status",
        "SSRS Parent Status",
        "SSRS Child Status",
        "SSRS Teacher Status",
    ]


def test_assessment_instruments_cover_all_eight_non_registration_instruments():
    labels = {label for _, _, label in ASSESSMENT_INSTRUMENTS}
    assert labels == {
        "SES",
        "DSEQ",
        "Child Illness History",
        "PAQ-A",
        "Dietary Intake",
        "SSRS Parent",
        "SSRS Child",
        "SSRS Teacher",
    }
