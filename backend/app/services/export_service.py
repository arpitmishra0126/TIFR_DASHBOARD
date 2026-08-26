"""Builds the "Active Cases" Excel export for the monthly newsletter.

Reads only from the already-normalized live REDCap data (RegistryChild +
raw records + choice maps) produced by LiveDashboardService — no new
REDCap fields are invented here, and no data is persisted. Field selection
follows the 2026-08-26 live-metadata field audit and the 2026-08-26
refinement pass: SES, DSEQ, Child Illness History, PAQ-A and Dietary
Intake contribute real acquired/derived columns; SSRS Parent/Child/Teacher
contribute per-child derived summaries (items answered + mean
frequency/importance rating) computed from their raw rating items rather
than a full raw-item dump, to keep the sheet readable — see
`ACTIVE_CASES_FIELD_SPECS` and the Data Dictionary sheet for exactly what
is exported and why. Free-text/notes fields, redundant identifier
duplicates, and caste/caste-category are intentionally excluded.
"""
from collections import Counter
from datetime import date, datetime
from io import BytesIO, StringIO
from statistics import mean
from typing import Callable, NamedTuple
import csv

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.ingestion.choice_maps import ChoiceMap
from app.ingestion.live_field_map import (
    CORE_BATTERY_COMPLETE_FIELDS,
    CORE_BATTERY_INSTRUMENTS,
    REGISTRATION_COMPLETE_FIELD,
    SSRS_CHILD_COMPLETE_FIELD,
    SSRS_CHILD_FREQ_FIELDS,
    SSRS_CHILD_IMP_FIELDS,
    SSRS_PARENT_FREQ_FIELDS,
    SSRS_PARENT_IMP_FIELDS,
    SSRS_TEACHER_COMPLETE_FIELD,
    SSRS_TEACHER_FREQ_FIELDS,
    SSRS_TEACHER_IMP_FIELDS,
)
from app.ingestion.normalize import parse_complete_flag, parse_date, parse_float, parse_int
from app.schemas.dashboard import RegistryChild
from app.services.module_analytics import (
    CHH_GENERAL_FLAGS,
    CHH_NAMED_CONDITIONS,
    DSEQ_YES_NO_ITEMS,
    INCOME_BUCKET_EDGES,
    INCOME_BUCKET_LABELS,
    PAQA_SCORE_BUCKET_EDGES,
    PAQA_SCORE_BUCKET_LABELS,
)
from app.services.module_analytics import bucket_counts as _bucket_counts
from app.services.module_analytics import category_counts as _category_counts
from app.services.module_analytics import complete_count as _complete_count
from app.services.module_analytics import numeric_values as _numeric_values
from app.services.module_analytics import resolve_value as _resolve_value
from app.services.module_analytics import yes_count as _yes_count

# All 8 non-registration instruments in PID 196, in newsletter order.
ASSESSMENT_INSTRUMENTS: tuple[tuple[str, str, str], ...] = (
    *CORE_BATTERY_INSTRUMENTS,
    ("ssrs_child", SSRS_CHILD_COMPLETE_FIELD, "SSRS Child"),
    ("ssrs_teacher", SSRS_TEACHER_COMPLETE_FIELD, "SSRS Teacher"),
)

_HEADER_FILL = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_GROUP_FILL = PatternFill(start_color="FF2F5597", end_color="FF2F5597", fill_type="solid")
_GROUP_FONT = Font(color="FFFFFFFF", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=13)
_SECTION_FONT = Font(bold=True, size=11)
_COMPLETE_FILL = PatternFill(start_color="FFE6F6E6", end_color="FFE6F6E6", fill_type="solid")
_NOT_COMPLETE_FILL = PatternFill(start_color="FFEEF0F4", end_color="FFEEF0F4", fill_type="solid")

_DEFAULT_MISSING_NOTE = "Blank = not collected (instrument not completed for this child)."


def _is_active(child: RegistryChild) -> bool:
    """A child is treated as an active case unless explicitly marked Dead.

    baby_status in the live project is a binary Live/Dead radio field
    (app.ingestion.live_field_map.REGISTRY_STATUS); a blank/unset status is
    treated as active since REDCap does not default new records to Dead.
    This definition is unchanged from the original export and from the
    live dashboard's own registry status field.
    """
    status = (child.child_status or "").strip().lower()
    return status != "dead"


class FieldSpec(NamedTuple):
    header: str
    redcap_field: str | None
    form: str
    kind: str  # "Identifier" | "Acquired" | "Derived" | "Status"
    redcap_type: str | None
    description: str
    missing_note: str
    group: str
    number_format: str | None
    value_fn: Callable[[RegistryChild, dict, dict[str, ChoiceMap]], object]


def _raw_spec(
    header: str, field: str, form: str, kind: str, redcap_type: str, group: str,
    description: str = "", missing_note: str = _DEFAULT_MISSING_NOTE,
) -> FieldSpec:
    def value_fn(_child: RegistryChild, record: dict, choice_maps: dict[str, ChoiceMap], field: str = field) -> object:
        value = _resolve_value(field, record.get(field), choice_maps)
        return value or None

    return FieldSpec(header, field, form, kind, redcap_type, description, missing_note, group, None, value_fn)


def _numeric_spec(
    header: str, field: str, form: str, kind: str, redcap_type: str, group: str,
    parser: Callable[[str | None], float | int | None], number_format: str,
    description: str = "", missing_note: str = _DEFAULT_MISSING_NOTE,
) -> FieldSpec:
    def value_fn(_child: RegistryChild, record: dict, _cm: dict[str, ChoiceMap], field: str = field, parser: Callable = parser) -> object:
        return parser(record.get(field))

    return FieldSpec(header, field, form, kind, redcap_type, description, missing_note, group, number_format, value_fn)


_GROUP_A = "A. Registration & Demographics"
_GROUP_B = "B. SES"
_GROUP_C = "C. DSEQ"
_GROUP_D = "D. Child Illness History"
_GROUP_E = "E. PAQ-A"
_GROUP_F = "F. Dietary Intake"
_GROUP_G = "G. SSRS Parent"
_GROUP_H = "H. SSRS Child"
_GROUP_I = "I. SSRS Teacher"
_GROUP_J = "J. Assessment / Progression Status"

_REGISTRY_SPECS: list[FieldSpec] = [
    FieldSpec(
        "Child ID", "child_id", "Registration", "Identifier", "text",
        "Primary record identifier.", "Never blank for an active case.", _GROUP_A, None,
        lambda c, r, cm: c.redcap_child_id,
    ),
    FieldSpec(
        "Sex", "baby_gender", "Registration", "Acquired", "dropdown",
        "Resolved to REDCap choice label.", "Blank = not recorded.", _GROUP_A, None,
        lambda c, r, cm: c.sex or None,
    ),
    FieldSpec(
        "Date of Birth", "child_dob", "Registration", "Acquired", "text",
        "", "Blank = not recorded.", _GROUP_A, "yyyy-mm-dd",
        lambda c, r, cm: parse_date(c.dob),
    ),
    FieldSpec(
        "Age (Years)", None, "Registration", "Derived", None,
        "Computed by the dashboard from Date of Birth as of export generation time.",
        "Blank = Date of Birth not recorded.", _GROUP_A, "0",
        lambda c, r, cm: c.age_years,
    ),
    FieldSpec(
        "Village", "village_name", "Registration", "Acquired", "text",
        "", "Blank = not recorded.", _GROUP_A, None, lambda c, r, cm: c.village or None,
    ),
    FieldSpec(
        "Child Status", "baby_status", "Registration", "Acquired", "radio",
        "Resolved to REDCap choice label (Live/Dead).", "Blank = not recorded.", _GROUP_A, None,
        lambda c, r, cm: c.child_status or None,
    ),
    FieldSpec(
        "Visit Date", "visit_date", "Registration", "Acquired", "text",
        "", "Blank = not recorded.", _GROUP_A, "yyyy-mm-dd",
        lambda c, r, cm: parse_date(c.visit_date),
    ),
]

_SES_MAPPED_SPECS: list[FieldSpec] = [
    _numeric_spec(
        "Udai Pareek SES Score", "scr_pareek_total", "SES", "Derived", "calc", _GROUP_B, parse_int, "0",
        "REDCap calculated field: cumulative score of Udai Pareek items P1-P9.",
        "Blank = SES questionnaire not completed for this child.",
    ),
    _numeric_spec(
        "Udai Pareek SES Category", "scr_pareek_category", "SES", "Derived", "calc", _GROUP_B, parse_int, "0",
        "REDCap calculated field. Numeric category code 1-5; REDCap does not expose a text label for calc fields.",
        "Blank = SES questionnaire not completed for this child.",
    ),
    _numeric_spec(
        "BG Prasad SES Category", "scr_prasad_category", "SES", "Derived", "calc", _GROUP_B, parse_int, "0",
        "REDCap calculated field. Numeric category code 1-5; REDCap does not expose a text label for calc fields.",
        "Blank = SES questionnaire not completed for this child.",
    ),
    _numeric_spec(
        "Per Capita Income", "scr_pci", "SES", "Derived", "calc", _GROUP_B, parse_float, "#,##0",
        "REDCap calculated field: monthly household income divided by household size.",
        "Blank = SES questionnaire not completed for this child.",
    ),
    _numeric_spec(
        "Household Size", "scr_bg_members", "SES", "Acquired", "text", _GROUP_B, parse_int, "0",
        "Raw household member count (field B2).",
        "Blank = SES questionnaire not completed for this child.",
    ),
    _numeric_spec(
        "Monthly Household Income (INR)", "scr_bg_income", "SES", "Acquired", "text", _GROUP_B, parse_float, "#,##0",
        "Raw total monthly family income (field B1); Per Capita Income above is the REDCap-derived figure.",
        "Blank = SES questionnaire not completed for this child.",
    ),
]

_SES_NEW_SPECS: list[FieldSpec] = [
    _raw_spec("Pareek P2: Occupation of Head of Family", "scr_pareek_occupation", "SES", "Acquired", "radio", _GROUP_B),
    _raw_spec("Pareek P3: Social Participation", "scr_pareek_social", "SES", "Acquired", "radio", _GROUP_B),
    _raw_spec("Pareek P4: House Type", "scr_pareek_house", "SES", "Acquired", "radio", _GROUP_B),
    _raw_spec("Pareek P5: Draught Animals Owned", "scr_pareek_animals", "SES", "Acquired", "radio", _GROUP_B),
    _raw_spec("Pareek P6: Head of Family Education", "scr_pareek_education", "SES", "Acquired", "radio", _GROUP_B),
    _raw_spec("Pareek P7: Land Owned", "scr_pareek_land", "SES", "Acquired", "radio", _GROUP_B),
    _raw_spec("Pareek P8: Assets Owned", "scr_pareek_assets", "SES", "Acquired", "radio", _GROUP_B),
    _raw_spec(
        "Pareek P9: Household Size Bucket", "scr_pareek_family", "SES", "Acquired", "radio", _GROUP_B,
        "Two-bucket variant (Up to 5 / More than 5); Household Size above is the precise numeric count.",
    ),
]

_DSEQ_LABELS: tuple[tuple[str, str], ...] = (
    ("q1_tv_freq", "DSEQ Q1: TV Frequency (per week)"),
    ("q2_tv_school", "DSEQ Q2: TV Time on School Day"),
    ("q3_tv_holiday", "DSEQ Q3: TV Time on Holiday"),
    ("q4_phone_freq", "DSEQ Q4: Smartphone/Tablet Frequency (per week)"),
    ("q5_phone_school", "DSEQ Q5: Smartphone/Tablet Time on School Day"),
    ("q6_phone_holiday", "DSEQ Q6: Smartphone/Tablet Time on Holiday"),
    ("q7_laptop_freq", "DSEQ Q7: Laptop/Computer Frequency (per week)"),
    ("q8_supervision", "DSEQ Q8: Adult Supervision During Screen Use"),
    ("q9_household_rules", "DSEQ Q9: Household Screen-Use Rules (Y/N)"),
    ("q10_total_screen_time", "DSEQ Q10: Average Total Daily Screen Time"),
    ("q11_outdoor_school", "DSEQ Q11: Outdoor Time on School Day"),
    ("q12_outdoor_holiday", "DSEQ Q12: Outdoor Time on Holiday"),
    ("q13_main_use", "DSEQ Q13: Main Use of Screen Devices"),
    ("q14_school_use", "DSEQ Q14: Uses Screens for School/Homework (Y/N)"),
    ("q15_entertainment_use", "DSEQ Q15: Uses Screens Mainly for Entertainment (Y/N)"),
)
_DSEQ_NEW_SPECS = [_raw_spec(label, field, "DSEQ", "Acquired", "radio", _GROUP_C) for field, label in _DSEQ_LABELS]

_CHH_LABELS: tuple[tuple[str, str], ...] = (
    ("chh_illness_current", "CHH: Currently Ill (Y/N)"),
    ("chh_unwell_7days", "CHH: Unwell in Past 7 Days (Y/N)"),
    ("chh_illness_3mo", "CHH: Illness Requiring Care in Past 3 Months (Y/N)"),
    ("chh_illness_3mo_freq", "CHH: Illness Frequency in Past 3 Months"),
    ("chh_missed_school_3mo", "CHH: Missed School Due to Illness (3mo)"),
    ("chh_chronic_condition", "CHH: Diagnosed Chronic Condition (Y/N)"),
    ("chh_q8_asthma", "CHH: Asthma (Y/N)"),
    ("chh_q8_heart", "CHH: Heart Disease (Y/N)"),
    ("chh_q8_tb", "CHH: Tuberculosis (Y/N)"),
    ("chh_q8_diabetes", "CHH: Diabetes (Y/N)"),
    ("chh_q8_thyroid", "CHH: Thyroid Disorder (Y/N)"),
    ("chh_q8_anaemia", "CHH: Anaemia (Y/N)"),
    ("chh_q8_malnutrition", "CHH: Severe Malnutrition (Y/N)"),
    ("chh_q8_kidney", "CHH: Kidney Disease (Y/N)"),
    ("chh_q8_liver", "CHH: Liver Disease (Y/N)"),
    ("chh_q8_infections", "CHH: Recurrent Serious Infections (Y/N)"),
    ("chh_q8_other", "CHH: Other Long-Term Condition (Y/N)"),
    ("chh_seizures", "CHH: Seizures/Fits/Convulsions (Y/N)"),
    ("chh_loc_fainting", "CHH: Unexplained Loss of Consciousness (Y/N)"),
    ("chh_cns_infection", "CHH: Meningitis/Encephalitis History (Y/N)"),
    ("chh_head_injury", "CHH: Significant Head Injury (Y/N)"),
    ("chh_vision_difficulty", "CHH: Vision Difficulty (Y/N)"),
    ("chh_uses_glasses", "CHH: Uses Glasses/Spectacles (Y/N)"),
    ("chh_hearing_difficulty", "CHH: Hearing Difficulty (Y/N)"),
    ("chh_ear_infection", "CHH: Recurrent Ear Infections (Y/N)"),
    ("chh_dev_diagnosis", "CHH: Developmental/Learning Diagnosis (Y/N)"),
    ("chh_hospitalised", "CHH: Ever Hospitalised Overnight (Y/N)"),
    ("chh_surgery", "CHH: Ever Had Surgery (Y/N)"),
    ("chh_medicine_current", "CHH: Currently on Regular Medicine (Y/N)"),
    ("chh_allergy", "CHH: Known Allergy (Y/N)"),
    ("chh_health_rating", "CHH: Overall Health Rating (vs. same-age children)"),
    ("chh_fit_for_assessment", "CHH: Fit for Today's Assessment"),
    ("chh_health_affects_today", "CHH: Health Condition May Affect Today's Assessment (Y/N)"),
    ("chh_assessor_decision", "CHH: Assessor Decision on Proceeding"),
)
_CHH_NEW_SPECS = [_raw_spec(label, field, "Child Illness History", "Acquired", "radio", _GROUP_D) for field, label in _CHH_LABELS]

_PAQA_NEW_SPECS: list[FieldSpec] = [
    _numeric_spec(
        "PAQ-A Item 1 Composite Score", "paq_item1_score", "PAQ-A", "Derived", "calc", _GROUP_E,
        lambda v: round(f, 2) if (f := parse_float(v)) is not None else None, "0.00",
        "REDCap calculated field: mean of the spare-time activity checklist.",
        "Blank = PAQ-A not completed for this child.",
    ),
    _numeric_spec(
        "PAQ-A Item 8 Composite Score", "paq_item8_score", "PAQ-A", "Derived", "calc", _GROUP_E,
        lambda v: round(f, 2) if (f := parse_float(v)) is not None else None, "0.00",
        "REDCap calculated field: mean of daily activity ratings, Monday-Sunday.",
        "Blank = PAQ-A not completed for this child.",
    ),
    _numeric_spec(
        "PAQ-A Total Score", "paq_total_score", "PAQ-A", "Derived", "calc", _GROUP_E,
        lambda v: round(f, 2) if (f := parse_float(v)) is not None else None, "0.00",
        "REDCap calculated field: mean of items 1-8 (excludes item 9, the illness-day check).",
        "Blank = PAQ-A not completed for this child.",
    ),
]

_DIETARY_LABELS: tuple[tuple[str, str], ...] = (
    ("die_grains_freq", "Diet Frequency: Grains / Roots / Tubers"),
    ("die_pulses_freq", "Diet Frequency: Pulses (Beans/Peas/Lentils)"),
    ("die_nuts_seeds_freq", "Diet Frequency: Nuts and Seeds"),
    ("die_dairy_freq", "Diet Frequency: Dairy (Milk/Yogurt/Cheese)"),
    ("die_flesh_freq", "Diet Frequency: Flesh Foods (Meat/Fish/Poultry)"),
    ("die_eggs_freq", "Diet Frequency: Eggs"),
    ("die_dgl_veg_freq", "Diet Frequency: Dark Green Leafy Vegetables"),
    ("die_vita_fv_freq", "Diet Frequency: Vitamin-A Rich Fruits/Vegetables"),
    ("die_other_veg_freq", "Diet Frequency: Other Vegetables"),
    ("die_other_fruits_freq", "Diet Frequency: Other Fruits"),
)
_DIETARY_NEW_SPECS = [_raw_spec(label, field, "Dietary Intake", "Acquired", "radio", _GROUP_F) for field, label in _DIETARY_LABELS]


def _ssrs_group_specs(
    freq_fields: tuple[str, ...], imp_fields: tuple[str, ...], group_label: str, form_label: str,
) -> list[FieldSpec]:
    """Per-child derived summary for one SSRS instrument: how many rating
    items were answered, and the mean of each rating scale. Raw items are
    not individually exported (would add ~90 columns per instrument); the
    Data Dictionary sheet documents the underlying REDCap field names.
    Computed identically for Parent/Child/Teacher so Teacher (0/212 live
    completions today) is handled by the same dynamic logic, not a special
    case — it will simply show 0 answered / blank averages until real
    Teacher data exists in REDCap.
    """
    n_freq = len(freq_fields)
    n_imp = len(imp_fields)

    def items_answered_fn(_c: RegistryChild, r: dict, _cm: dict) -> str:
        answered = sum(1 for f in freq_fields if (r.get(f) or "").strip() != "")
        return f"{answered}/{n_freq}"

    def avg_fn(fields: tuple[str, ...]) -> Callable[[RegistryChild, dict, dict], float | None]:
        def value_fn(_c: RegistryChild, r: dict, _cm: dict) -> float | None:
            values = [v for f in fields if (v := parse_float(r.get(f))) is not None]
            return round(mean(values), 2) if values else None

        return value_fn

    return [
        FieldSpec(
            f"{form_label}: Items Answered", None, form_label, "Derived", "radio (aggregated)",
            f"Count of the {n_freq} frequency-rating items answered, out of {n_freq}.",
            "0/N = instrument not started for this child.", group_label, None, items_answered_fn,
        ),
        FieldSpec(
            f"{form_label}: Avg Frequency Rating", None, form_label, "Derived", "radio (aggregated)",
            "Mean of the answered frequency-rating items (REDCap codes 0=Never, 1=Sometimes, 2=Very Often); "
            "computed by the export, not a validated SSRS composite score.",
            "Blank = no frequency items answered.", group_label, "0.00", avg_fn(freq_fields),
        ),
        FieldSpec(
            f"{form_label}: Avg Importance Rating", None, form_label, "Derived", "radio (aggregated)",
            "Mean of the answered importance-rating items (REDCap codes 0=Not Important, 1=Important, ...); "
            "computed by the export, not a validated SSRS composite score.",
            "Blank = no importance items answered.", group_label, "0.00", avg_fn(imp_fields),
        ),
    ]


_SSRS_PARENT_SPECS = _ssrs_group_specs(SSRS_PARENT_FREQ_FIELDS, SSRS_PARENT_IMP_FIELDS, _GROUP_G, "SSRS Parent")
_SSRS_CHILD_SPECS = _ssrs_group_specs(SSRS_CHILD_FREQ_FIELDS, SSRS_CHILD_IMP_FIELDS, _GROUP_H, "SSRS Child")
_SSRS_TEACHER_SPECS = _ssrs_group_specs(SSRS_TEACHER_FREQ_FIELDS, SSRS_TEACHER_IMP_FIELDS, _GROUP_I, "SSRS Teacher")


def _progression_stage_fn(_c: RegistryChild, r: dict, _cm: dict) -> str:
    if not all(parse_complete_flag(r.get(f)) for f in CORE_BATTERY_COMPLETE_FIELDS):
        return "Registered"
    if not parse_complete_flag(r.get(SSRS_CHILD_COMPLETE_FIELD)):
        return "Core Assessment Battery"
    if not parse_complete_flag(r.get(SSRS_TEACHER_COMPLETE_FIELD)):
        return "SSRS Child"
    return "SSRS Teacher"


def _complete_status_fn(field: str) -> Callable[[RegistryChild, dict, dict], str]:
    def value_fn(_c: RegistryChild, r: dict, _cm: dict, field: str = field) -> str:
        return "Complete" if parse_complete_flag(r.get(field)) else "Not Complete"

    return value_fn


def _core_battery_status_fn(_c: RegistryChild, r: dict, _cm: dict) -> str:
    return "Complete" if all(parse_complete_flag(r.get(f)) for f in CORE_BATTERY_COMPLETE_FIELDS) else "Not Complete"


_PROGRESSION_SPECS: list[FieldSpec] = [
    FieldSpec(
        "Registration Complete", REGISTRATION_COMPLETE_FIELD, "Registration", "Status", "text (system)",
        "REDCap instrument completion flag (0/1/2).", "N/A — always Complete or Not Complete.", _GROUP_J, None,
        lambda c, r, cm: "Complete" if c.registration_complete else "Not Complete",
    ),
    *[
        FieldSpec(
            f"{label} Complete", field, label, "Status", "text (system)",
            "REDCap instrument completion flag (0/1/2).", "N/A — always Complete or Not Complete.", _GROUP_J, None,
            _complete_status_fn(field),
        )
        for _, field, label in ASSESSMENT_INSTRUMENTS
    ],
    FieldSpec(
        "Core Assessment Battery", None, "Multiple (SES, DSEQ, Child Illness History, PAQ-A, Dietary Intake, SSRS Parent)",
        "Status", None, "Complete only when all six core-battery instruments are complete for this child.",
        "N/A — always Complete or Not Complete.", _GROUP_J, None, _core_battery_status_fn,
    ),
    FieldSpec(
        "Overall Progression Stage", None, "Multiple", "Status", None,
        "Highest pipeline stage reached: Registered -> Core Assessment Battery -> SSRS Child -> SSRS Teacher "
        "(same cumulative definition as the dashboard's Assessment Progress module).",
        "N/A — always populated.", _GROUP_J, None, _progression_stage_fn,
    ),
]

ACTIVE_CASES_FIELD_SPECS: list[FieldSpec] = (
    _REGISTRY_SPECS
    + _SES_MAPPED_SPECS
    + _SES_NEW_SPECS
    + _DSEQ_NEW_SPECS
    + _CHH_NEW_SPECS
    + _PAQA_NEW_SPECS
    + _DIETARY_NEW_SPECS
    + _SSRS_PARENT_SPECS
    + _SSRS_CHILD_SPECS
    + _SSRS_TEACHER_SPECS
    + _PROGRESSION_SPECS
)

_EXCLUDED_FIELDS_NOTE: tuple[tuple[str, str], ...] = (
    ("scr_pareek_caste", "Caste (Udai Pareek P1) — excluded per data-sensitivity policy."),
    ("scr_caste_category", "Actual caste category (P1a) — excluded per data-sensitivity policy."),
    ("parent_child_id", "Redundant duplicate identifier on the SSRS Parent form — the Child ID column already covers this."),
    ("teacher_child_id", "Redundant duplicate identifier on the SSRS Teacher form — the Child ID column already covers this."),
    ("*_spec / *_comment / *_reason / *_remarks (all instruments)", "Free-text/clinical-narrative fields — excluded to avoid re-identification risk."),
    (
        "parent_complete / teacher_complete",
        "These are descriptive scoring-rubric text blocks on the SSRS forms, not real completion flags. True "
        "completion is read from the record-level ssrs_parent_complete/ssrs_teacher_complete fields.",
    ),
    (
        "p*_freq / p*_imp / c*_freq / c*_imp / t*_freq / t*_imp (raw SSRS items)",
        "92 (Parent) + 68 (Child) + 72 (Teacher) individual rating items are not exported one-per-column to keep "
        "the sheet readable — see the SSRS Parent/Child/Teacher 'Items Answered'/'Avg Frequency Rating'/'Avg "
        "Importance Rating' derived columns instead.",
    ),
)


def _style_header_row(sheet: Worksheet, row: int, num_columns: int) -> None:
    for col in range(1, num_columns + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _autosize_columns(sheet: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = max(width, 10)


def _complete_label(record: dict, field: str) -> str:
    """Used only by the CSV export, which keeps its original wording."""
    return "Complete" if parse_complete_flag(record.get(field)) else "Incomplete"


def _status_label(record: dict, field: str) -> str:
    """Used by the Excel Assessment Status sheet: 'Complete' / 'Not Complete'."""
    return "Complete" if parse_complete_flag(record.get(field)) else "Not Complete"


def _record_by_child_id(records: list[dict]) -> dict[str, dict]:
    by_id: dict[str, dict] = {}
    for record in records:
        child_id = (record.get("child_id") or "").strip()
        if child_id:
            by_id[child_id] = record
    return by_id


def _build_active_cases_sheet(
    sheet: Worksheet,
    active_children: list[RegistryChild],
    records_by_id: dict[str, dict],
    choice_maps: dict[str, ChoiceMap],
) -> None:
    specs = ACTIVE_CASES_FIELD_SPECS
    n = len(specs)

    # Row 1: merged group headers (A. Registration & Demographics, B. SES, ...).
    col = 1
    while col <= n:
        group = specs[col - 1].group
        start = col
        while col <= n and specs[col - 1].group == group:
            col += 1
        end = col - 1
        if end > start:
            sheet.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        for c in range(start, end + 1):
            cell = sheet.cell(row=1, column=c)
            cell.fill = _GROUP_FILL
            cell.font = _GROUP_FONT
        header_cell = sheet.cell(row=1, column=start, value=group)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: individual column headers.
    for idx, spec in enumerate(specs, start=1):
        cell = sheet.cell(row=2, column=idx, value=spec.header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[2].height = 45

    # Data rows from row 3.
    for child in active_children:
        record = records_by_id.get(child.redcap_child_id, {})
        row_idx = sheet.max_row + 1
        for col_idx, spec in enumerate(specs, start=1):
            value = spec.value_fn(child, record, choice_maps)
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if spec.number_format:
                cell.number_format = spec.number_format

    last_row = sheet.max_row
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(n)}{last_row}"
    _autosize_columns(sheet, [max(14, min(30, len(spec.header))) for spec in specs])


def _build_assessment_status_sheet(sheet: Worksheet, active_children: list[RegistryChild], records_by_id: dict[str, dict]) -> None:
    headers = [
        "Child ID", "Registration", *(label for _, _, label in ASSESSMENT_INSTRUMENTS),
        "Core Assessment Battery", "Overall Progression Stage",
    ]
    sheet.append(headers)
    _style_header_row(sheet, 1, len(headers))

    for child in active_children:
        record = records_by_id.get(child.redcap_child_id, {})
        row = [
            child.redcap_child_id,
            "Complete" if child.registration_complete else "Not Complete",
            *(_status_label(record, field) for _, field, _ in ASSESSMENT_INSTRUMENTS),
            _core_battery_status_fn(child, record, {}),
            _progression_stage_fn(child, record, {}),
        ]
        sheet.append(row)

    # Light fill so "Complete" rows are visually distinct at a glance.
    for row_cells in sheet.iter_rows(min_row=2, min_col=2, max_col=len(ASSESSMENT_INSTRUMENTS) + 3):
        for cell in row_cells:
            if cell.value == "Complete":
                cell.fill = _COMPLETE_FILL
            elif cell.value == "Not Complete":
                cell.fill = _NOT_COMPLETE_FILL

    _autosize_columns(sheet, [14, 14, *([18] * len(ASSESSMENT_INSTRUMENTS)), 20, 22])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"


def _build_data_dictionary_sheet(sheet: Worksheet) -> None:
    headers = ["Export Column", "REDCap Variable", "Instrument", "Description", "Data Type", "Acquired vs Derived", "Missing-Data Interpretation"]
    sheet.append(headers)
    _style_header_row(sheet, 1, len(headers))

    for spec in ACTIVE_CASES_FIELD_SPECS:
        sheet.append([
            spec.header, spec.redcap_field or "(computed)", spec.form, spec.description or "",
            spec.redcap_type or "", spec.kind, spec.missing_note,
        ])

    sheet.append([])
    sheet.append(["Assessment Status sheet — one row per active child:"])
    sheet.cell(row=sheet.max_row, column=1).font = _SECTION_FONT
    status_header_row = sheet.max_row + 1
    sheet.append(headers)
    _style_header_row(sheet, status_header_row, len(headers))
    sheet.append(["Registration", REGISTRATION_COMPLETE_FIELD, "Registration", "Instrument completion status.", "text (system)", "Status", "N/A"])
    for _, field, label in ASSESSMENT_INSTRUMENTS:
        sheet.append([f"{label}", field, label, "Instrument completion status.", "text (system)", "Status", "N/A"])
    sheet.append(["Core Assessment Battery", None, "Multiple", "Complete only when all six core-battery instruments are complete.", "", "Status", "N/A"])
    sheet.append(["Overall Progression Stage", None, "Multiple", "Highest pipeline stage reached for this child.", "", "Status", "N/A"])

    sheet.append([])
    sheet.append([
        "Summary sheet contains only aggregated/derived statistics computed from the Active Cases and "
        "Assessment Status sheets (counts, percentages, distributions) — no additional REDCap fields are read."
    ])

    sheet.append([])
    sheet.append(["Intentionally excluded fields (per data-sensitivity / free-text / readability policy):"])
    sheet.cell(row=sheet.max_row, column=1).font = _SECTION_FONT
    for field, reason in _EXCLUDED_FIELDS_NOTE:
        sheet.append([field, "", "", reason, "", "Excluded", ""])

    _autosize_columns(sheet, [34, 22, 30, 70, 16, 16, 44])
    sheet.freeze_panes = "A2"


def _write_table(
    sheet: Worksheet,
    row: int,
    title: str,
    col_headers: list[str],
    data_rows: list[tuple],
) -> tuple[int, int, int, int]:
    """Writes a titled table starting at (row, column A). Returns
    (next_free_row, header_row, first_data_row, last_data_row)."""
    sheet.cell(row=row, column=1, value=title).font = _SECTION_FONT
    row += 1
    header_row = row
    for col, header in enumerate(col_headers, start=1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
    row += 1
    first_data_row = row
    for values in data_rows:
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row, column=col, value=value)
        row += 1
    last_data_row = row - 1
    return row + 1, header_row, first_data_row, last_data_row


def _add_bar_chart(
    sheet: Worksheet,
    anchor: str,
    title: str,
    header_row: int,
    first_row: int,
    last_row: int,
    y_title: str = "Children",
    show_data_labels: bool = True,
) -> None:
    """Adds a column chart over a 2-column (category, value) table.
    No-ops if the table is empty or every value is zero — charts are never
    drawn for variables with insufficient/no data. All charts use a fixed
    size and styling for a consistent, report-like appearance.
    """
    if last_row < first_row:
        return
    total = sum((sheet.cell(row=r, column=2).value or 0) for r in range(first_row, last_row + 1))
    if not total:
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = y_title
    chart.style = 10
    chart.legend = None
    chart.height = 8
    chart.width = 16
    chart.gapWidth = 60

    data = Reference(sheet, min_col=2, min_row=header_row, max_row=last_row)
    cats = Reference(sheet, min_col=1, min_row=first_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if show_data_labels:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
    sheet.add_chart(chart, anchor)


# --- Pure aggregation helpers (export-local; intentionally independent of
# LiveDashboardService's internals to avoid a circular import). ---

_AGE_BUCKETS: tuple[tuple[str, int, int], ...] = (("0-4", 0, 4), ("5-9", 5, 9), ("10-14", 10, 14), ("15+", 15, 200))


def _age_distribution(children: list[RegistryChild]) -> list[tuple[str, int]]:
    buckets = {label: 0 for label, _, _ in _AGE_BUCKETS}
    unknown = 0
    for child in children:
        if child.age_years is None:
            unknown += 1
            continue
        for label, low, high in _AGE_BUCKETS:
            if low <= child.age_years <= high:
                buckets[label] += 1
                break
    rows = [(label, buckets[label]) for label, _, _ in _AGE_BUCKETS]
    if unknown:
        rows.append(("Unknown", unknown))
    return rows


def _sex_counts(children: list[RegistryChild]) -> list[tuple[str, int]]:
    male = sum(1 for c in children if (c.sex or "").strip().lower() == "male")
    female = sum(1 for c in children if (c.sex or "").strip().lower() == "female")
    unknown = len(children) - male - female
    rows = [("Male", male), ("Female", female)]
    if unknown:
        rows.append(("Unknown", unknown))
    return rows


def _village_counts(children: list[RegistryChild]) -> list[tuple[str, int]]:
    counts = Counter(c.village for c in children if c.village)
    return sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))


def _numeric_summary_line(label: str, values: list[float], total: int) -> str:
    if not values:
        return f"{label}: No data acquired (0/{total})."
    return (
        f"{label}: n={len(values)}/{total} acquired ({_percent(len(values), total)}), "
        f"mean={round(mean(values), 1)}, min={min(values)}, max={max(values)}"
    )


def _percent(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return "0.0%"
    return f"{round((numerator / denominator) * 100, 1)}%"


def _unique_ids_with_complete(active_records: list[dict], field: str) -> set[str]:
    return {cid for r in active_records if (cid := (r.get("child_id") or "").strip()) and parse_complete_flag(r.get(field))}


def _core_battery_ids(active_records: list[dict]) -> set[str]:
    per_field_ids = [_unique_ids_with_complete(active_records, field) for field in CORE_BATTERY_COMPLETE_FIELDS]
    if not per_field_ids:
        return set()
    result = per_field_ids[0]
    for ids in per_field_ids[1:]:
        result &= ids
    return result


def _write_count_row(sheet: Worksheet, row: int, label: str, count: int, total: int) -> int:
    sheet.cell(row=row, column=1, value=label)
    sheet.cell(row=row, column=2, value=count)
    sheet.cell(row=row, column=3, value=total)
    pct_cell = sheet.cell(row=row, column=4, value=(count / total) if total else 0)
    pct_cell.number_format = "0.0%"
    return row + 1


def _write_coverage_note(sheet: Worksheet, row: int, text: str) -> int:
    sheet.cell(row=row, column=1, value=text)
    return row + 1


def _ssrs_items_answered(active_records: list[dict], freq_fields: tuple[str, ...]) -> int:
    return sum(1 for r in active_records if any((r.get(f) or "").strip() != "" for f in freq_fields))


def _build_summary_sheet(
    sheet: Worksheet,
    all_children: list[RegistryChild],
    active_children: list[RegistryChild],
    active_records: list[dict],
    choice_maps: dict[str, ChoiceMap],
    generated_at: datetime,
) -> None:
    sheet.column_dimensions["A"].width = 46
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12

    total_registered = len(all_children)
    total_active = len(active_children)
    chart_row = 2  # running anchor row for charts, stacked in column F

    def next_chart_anchor() -> str:
        nonlocal chart_row
        anchor = f"F{chart_row}"
        chart_row += 19
        return anchor

    row = 1
    sheet.cell(row=row, column=1, value="ICMR Neurodevelopment Study — Active Cases Analysis").font = _TITLE_FONT
    row += 1
    sheet.cell(row=row, column=1, value=f"Generated: {generated_at.strftime('%Y-%m-%d %H:%M:%S')} (live REDCap data)")
    row += 1
    sheet.cell(row=row, column=1, value="Active case definition: registration status is not 'Dead' (registration_form.baby_status). Unchanged from prior export versions.")
    row += 2

    # ================= POPULATION =================
    sheet.cell(row=row, column=1, value="POPULATION").font = _TITLE_FONT
    row += 1
    header_row = row
    for col, h in enumerate(["Metric", "Count", "of Total", "Percent"], start=1):
        sheet.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1
    row = _write_count_row(sheet, row, "Total Registered (all statuses)", total_registered, total_registered)
    row = _write_count_row(sheet, row, "Total Active Cases", total_active, total_registered)
    row += 1

    row, hdr, first, last = _write_table(sheet, row, "Sex Distribution (Active Cases)", ["Sex", "Count"], _sex_counts(active_children))
    _add_bar_chart(sheet, next_chart_anchor(), "Sex Distribution", hdr, first, last)

    row, hdr, first, last = _write_table(sheet, row, "Age Distribution (Active Cases)", ["Age Bucket", "Count"], _age_distribution(active_children))
    _add_bar_chart(sheet, next_chart_anchor(), "Age Distribution", hdr, first, last)

    village_rows = _village_counts(active_children)
    row, hdr, first, last = _write_table(sheet, row, f"Village Distribution — All {len(village_rows)} Villages (Active Cases)", ["Village", "Count"], village_rows)
    top10_last = min(last, first + 9)
    _add_bar_chart(sheet, next_chart_anchor(), f"Top 10 Villages by Active Cases (of {len(village_rows)} total)", hdr, first, top10_last)

    # ================= SES =================
    sheet.cell(row=row, column=1, value="SES").font = _TITLE_FONT
    row += 1

    pareek_scores = _numeric_values(active_records, "scr_pareek_total", parse_float)
    row, hdr, first, last = _write_table(
        sheet, row, "Udai Pareek SES Category Distribution", ["Category (code)", "Count"],
        _category_counts(active_records, "scr_pareek_category", choice_maps),
    )
    row = _write_coverage_note(sheet, row, _numeric_summary_line("Udai Pareek score", pareek_scores, total_active))
    row += 1
    _add_bar_chart(sheet, next_chart_anchor(), "Udai Pareek Category Distribution", hdr, first, last)

    row, hdr, first, last = _write_table(
        sheet, row, "BG Prasad SES Category Distribution", ["Category (code)", "Count"],
        _category_counts(active_records, "scr_prasad_category", choice_maps),
    )
    _add_bar_chart(sheet, next_chart_anchor(), "BG Prasad Category Distribution", hdr, first, last)

    per_capita_values = _numeric_values(active_records, "scr_pci", parse_float)
    income_buckets = _bucket_counts(per_capita_values, INCOME_BUCKET_EDGES, INCOME_BUCKET_LABELS)
    row, hdr, first, last = _write_table(sheet, row, "Monthly Income — Per Capita Income Distribution (INR)", ["Range", "Count"], income_buckets)
    monthly_income_values = _numeric_values(active_records, "scr_bg_income", parse_float)
    row = _write_coverage_note(sheet, row, _numeric_summary_line("Per capita income", per_capita_values, total_active))
    row = _write_coverage_note(sheet, row, _numeric_summary_line("Raw monthly household income", monthly_income_values, total_active))
    row += 1
    _add_bar_chart(sheet, next_chart_anchor(), "Per Capita Income Distribution", hdr, first, last)

    household_sizes = _numeric_values(active_records, "scr_bg_members", lambda v: parse_int(v))
    size_counts = Counter(int(v) for v in household_sizes)
    size_rows = sorted(size_counts.items())
    row, hdr, first, last = _write_table(sheet, row, "Household Size Distribution", ["Household Size", "Count"], size_rows)
    _add_bar_chart(sheet, next_chart_anchor(), "Household Size Distribution", hdr, first, last)

    sheet.cell(row=row, column=1, value="SES — Additional Udai Pareek Items (P2-P9, tables only)").font = _SECTION_FONT
    row += 1
    for field, label in (
        ("scr_pareek_occupation", "P2: Occupation of Head of Family"),
        ("scr_pareek_social", "P3: Social Participation"),
        ("scr_pareek_house", "P4: House Type"),
        ("scr_pareek_animals", "P5: Draught Animals Owned"),
        ("scr_pareek_education", "P6: Head of Family Education"),
        ("scr_pareek_land", "P7: Land Owned"),
        ("scr_pareek_assets", "P8: Assets Owned"),
        ("scr_pareek_family", "P9: Household Size Bucket"),
    ):
        row, _, _, _ = _write_table(sheet, row, label, ["Response", "Count"], _category_counts(active_records, field, choice_maps))
    row += 1

    # ================= ASSESSMENT ACQUISITION =================
    sheet.cell(row=row, column=1, value="ASSESSMENT ACQUISITION").font = _TITLE_FONT
    row += 1
    registration_complete_count = sum(1 for c in active_children if c.registration_complete)
    instrument_rows = [("Registration", registration_complete_count)]
    for _, field, label in ASSESSMENT_INSTRUMENTS:
        instrument_rows.append((label, _complete_count(active_records, field)))
    row, hdr, first, last = _write_table(sheet, row, "Instrument-wise Acquired/Completed Counts", ["Instrument", "Completed (of Active)"], instrument_rows)
    # Add a percent column alongside for readability.
    for r in range(first, last + 1):
        count = sheet.cell(row=r, column=2).value or 0
        pct_cell = sheet.cell(row=r, column=3, value=(count / total_active) if total_active else 0)
        pct_cell.number_format = "0.0%"
    sheet.cell(row=hdr, column=3, value="% of Active").font = Font(bold=True)
    _add_bar_chart(sheet, next_chart_anchor(), "Instrument Acquisition (Completed Count)", hdr, first, last)

    # ================= ASSESSMENT PROGRESSION =================
    sheet.cell(row=row, column=1, value="ASSESSMENT PROGRESSION").font = _TITLE_FONT
    row += 1
    core_ids = _core_battery_ids(active_records)
    ssrs_child_ids = core_ids & _unique_ids_with_complete(active_records, SSRS_CHILD_COMPLETE_FIELD)
    ssrs_teacher_ids = ssrs_child_ids & _unique_ids_with_complete(active_records, SSRS_TEACHER_COMPLETE_FIELD)
    stage_counts = [total_active, len(core_ids), len(ssrs_child_ids), len(ssrs_teacher_ids)]
    stage_labels = ["Registered", "Core Assessment Battery", "SSRS Child", "SSRS Teacher"]
    progression_rows = list(zip(stage_labels, stage_counts))
    row, hdr, first, last = _write_table(sheet, row, "Assessment Progression (Cumulative)", ["Stage", "Active Children"], progression_rows)
    _add_bar_chart(sheet, next_chart_anchor(), "Assessment Progression", hdr, first, last)

    stage_pct_rows = [
        ("Registered -> Core Assessment Battery", _percent(stage_counts[1], stage_counts[0])),
        ("Core Assessment Battery -> SSRS Child", _percent(stage_counts[2], stage_counts[1]) if stage_counts[1] else "N/A (no cases reached this stage)"),
        ("SSRS Child -> SSRS Teacher", _percent(stage_counts[3], stage_counts[2]) if stage_counts[2] else "N/A (no cases reached this stage)"),
    ]
    row, _, _, _ = _write_table(sheet, row, "Stage-to-Stage Conversion", ["Transition", "% Advancing"], stage_pct_rows)
    row += 1

    # ================= DOMAIN ANALYSIS =================
    sheet.cell(row=row, column=1, value="DOMAIN ANALYSIS (where sufficient data exists)").font = _TITLE_FONT
    row += 1

    # --- DSEQ ---
    row, hdr, first, last = _write_table(
        sheet, row, "DSEQ — Average Total Daily Screen Time (Q10)", ["Response", "Count"],
        _category_counts(active_records, "q10_total_screen_time", choice_maps),
    )
    _add_bar_chart(sheet, next_chart_anchor(), "DSEQ: Total Daily Screen Time", hdr, first, last)
    dseq_yn_rows = [(label, _yes_count(active_records, field, choice_maps)) for field, label in DSEQ_YES_NO_ITEMS]
    row, _, _, _ = _write_table(sheet, row, "DSEQ — Selected Yes/No Items (of Active)", ["Item", "Yes Count"], dseq_yn_rows)

    # --- Child Illness History ---
    condition_rows = [(label, _yes_count(active_records, field, choice_maps)) for field, label in CHH_NAMED_CONDITIONS]
    row, hdr, first, last = _write_table(sheet, row, "Child Illness History — Named Conditions (Yes counts)", ["Condition", "Yes Count"], condition_rows)
    _add_bar_chart(sheet, next_chart_anchor(), "CHH: Named Conditions (Yes)", hdr, first, last)

    chh_general_rows = [(label, _yes_count(active_records, field, choice_maps)) for field, label in CHH_GENERAL_FLAGS]
    row, _, _, _ = _write_table(sheet, row, "Child Illness History — General Flags (Yes counts, of Active)", ["Flag", "Yes Count"], chh_general_rows)

    # --- PAQ-A ---
    item1 = _numeric_values(active_records, "paq_item1_score", parse_float)
    item8 = _numeric_values(active_records, "paq_item8_score", parse_float)
    total_scores = _numeric_values(active_records, "paq_total_score", parse_float)
    row = _write_coverage_note(sheet, row, _numeric_summary_line("PAQ-A Item 1 score", item1, total_active))
    row = _write_coverage_note(sheet, row, _numeric_summary_line("PAQ-A Item 8 score", item8, total_active))
    row = _write_coverage_note(sheet, row, _numeric_summary_line("PAQ-A Total score", total_scores, total_active))
    row += 1
    paqa_buckets = _bucket_counts(total_scores, PAQA_SCORE_BUCKET_EDGES, PAQA_SCORE_BUCKET_LABELS)
    row, hdr, first, last = _write_table(sheet, row, "PAQ-A Total Score Distribution", ["Range", "Count"], paqa_buckets)
    _add_bar_chart(sheet, next_chart_anchor(), "PAQ-A Total Score Distribution", hdr, first, last)

    # --- Dietary Intake ---
    daily_rows = []
    for field, label in _DIETARY_LABELS:
        # "Daily" = REDCap response code 1, consistent across all 10 dietary frequency items in this project.
        daily_count = sum(1 for r in active_records if (r.get(field) or "").strip() == "1")
        daily_rows.append((label.replace("Diet Frequency: ", ""), daily_count))
    row, hdr, first, last = _write_table(
        sheet, row, "Dietary Intake — Daily Consumption by Food Group (of Active)", ["Food Group", "Children Reporting Daily"], daily_rows,
    )
    _add_bar_chart(sheet, next_chart_anchor(), "Dietary: Daily Consumption by Food Group", hdr, first, last)

    # --- SSRS ---
    ssrs_rows = [
        ("SSRS Parent", _ssrs_items_answered(active_records, SSRS_PARENT_FREQ_FIELDS)),
        ("SSRS Child", _ssrs_items_answered(active_records, SSRS_CHILD_FREQ_FIELDS)),
        ("SSRS Teacher", _ssrs_items_answered(active_records, SSRS_TEACHER_FREQ_FIELDS)),
    ]
    row, hdr, first, last = _write_table(sheet, row, "SSRS — Children With Any Rating Items Answered (of Active)", ["Instrument", "Children"], ssrs_rows)
    _add_bar_chart(sheet, next_chart_anchor(), "SSRS: Children With Data", hdr, first, last)
    row = _write_coverage_note(
        sheet, row,
        "SSRS Teacher has 0 completed assessments live — its row above and its Active Cases columns will "
        "populate automatically once REDCap has real Teacher data; no value has been invented here.",
    )
    row += 1

    # ================= DATA COVERAGE =================
    sheet.cell(row=row, column=1, value="DATA COVERAGE").font = _TITLE_FONT
    row += 1
    row = _write_coverage_note(sheet, row, f"Total Active Cases: {total_active}")
    row += 1

    coverage_rows = []
    coverage_rows.append(("Registration", registration_complete_count, total_active))
    for _, field, label in ASSESSMENT_INSTRUMENTS:
        coverage_rows.append((label, _complete_count(active_records, field), total_active))

    high = [c for c in coverage_rows if c[2] and c[1] / c[2] >= 0.5]
    partial = [c for c in coverage_rows if c[2] and 0 < c[1] / c[2] < 0.5]
    none_ = [c for c in coverage_rows if c[1] == 0]

    def _coverage_table(title: str, rows: list[tuple[str, int, int]]) -> None:
        nonlocal row
        sheet.cell(row=row, column=1, value=title).font = _SECTION_FONT
        row += 1
        if not rows:
            row = _write_coverage_note(sheet, row, "(none)")
            row += 1
            return
        for label, count, total in rows:
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=count)
            sheet.cell(row=row, column=3, value=total)
            pct_cell = sheet.cell(row=row, column=4, value=(count / total) if total else 0)
            pct_cell.number_format = "0.0%"
            row += 1
        row += 1

    _coverage_table("High Coverage (>= 50% of Active Cases)", high)
    _coverage_table("Partially Acquired (> 0% and < 50% of Active Cases)", partial)
    _coverage_table("No Data Acquired Yet (0% of Active Cases)", none_)

    sheet.append([])
    sheet.append([
        "Note: SSRS Parent and SSRS Child now show real per-child derived summaries (Items Answered, Avg "
        "Frequency/Importance Rating) in the Active Cases sheet where data exists. SSRS Teacher's columns "
        "are computed the same way but show 0/blank today because 0 Teacher assessments are complete live."
    ])


def build_active_cases_workbook(
    all_children: list[RegistryChild],
    records: list[dict],
    choice_maps: dict[str, ChoiceMap],
    generated_at: datetime | None = None,
) -> bytes:
    """Build the Active Cases newsletter workbook as raw .xlsx bytes."""
    generated_at = generated_at or datetime.now()
    active_children = [c for c in all_children if _is_active(c)]
    records_by_id = _record_by_child_id(records)
    active_records = [records_by_id[c.redcap_child_id] for c in active_children if c.redcap_child_id in records_by_id]

    workbook = Workbook()

    active_sheet = workbook.active
    active_sheet.title = "Active Cases"
    _build_active_cases_sheet(active_sheet, active_children, records_by_id, choice_maps)

    status_sheet = workbook.create_sheet("Assessment Status")
    _build_assessment_status_sheet(status_sheet, active_children, records_by_id)

    summary_sheet = workbook.create_sheet("Summary")
    _build_summary_sheet(summary_sheet, all_children, active_children, active_records, choice_maps, generated_at)

    dictionary_sheet = workbook.create_sheet("Data Dictionary")
    _build_data_dictionary_sheet(dictionary_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_filename(as_of: date | None = None, extension: str = "xlsx") -> str:
    as_of = as_of or date.today()
    return f"ICMR_Active_Cases_{as_of.isoformat()}.{extension}"


def build_active_cases_csv(all_children: list[RegistryChild], records: list[dict]) -> str:
    """Build the Active Cases CSV export: one row per active child.

    Kept at its original (pre-audit-expansion) field set — Registration/SES
    core fields + per-instrument completion status — since the CSV export
    was not part of the Excel refinement requests. Unmapped-instrument
    columns are left blank (not the descriptive placeholder text used in
    the workbook).
    """
    active_children = [c for c in all_children if _is_active(c)]
    records_by_id = _record_by_child_id(records)

    buffer = StringIO()
    writer = csv.writer(buffer)

    headers = [
        "Child ID",
        "Sex",
        "Date of Birth",
        "Age (Years)",
        "Village",
        "Child Status",
        "Visit Date",
        "Registration Complete",
        "Udai Pareek SES Score",
        "Udai Pareek SES Category",
        "BG Prasad SES Category",
        "Per Capita Income",
        "Household Size",
        *(f"{label} Status" for _, _, label in ASSESSMENT_INSTRUMENTS),
    ]
    writer.writerow(headers)

    for child in active_children:
        record = records_by_id.get(child.redcap_child_id, {})
        per_capita_income = parse_float(record.get("scr_pci"))
        household_size = parse_int(record.get("scr_bg_members"))
        row = [
            child.redcap_child_id,
            child.sex or "",
            child.dob or "",
            child.age_years if child.age_years is not None else "",
            child.village or "",
            child.child_status or "",
            child.visit_date or "",
            "Yes" if child.registration_complete else "No",
            record.get("scr_pareek_total") or "",
            record.get("scr_pareek_category") or "",
            record.get("scr_prasad_category") or "",
            per_capita_income if per_capita_income is not None else "",
            household_size if household_size is not None else "",
            *(_complete_label(record, field) for _, field, _ in ASSESSMENT_INSTRUMENTS),
        ]
        writer.writerow(row)

    return buffer.getvalue()
